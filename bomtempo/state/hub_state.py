"""
HubState — State for Hub de Operações sub-modules:
  • Cronograma: CRUD de atividades, edição inline, gestão de dependências
  • Auditoria: Bolsões de imagens (Equipe, Falhas, Ferramentas, Gerais) com lightbox
  • Timeline: Log de eventos/registros do projeto, comentários, @mentions

Supabase tables required (see schema below):
  - hub_atividades: id, contrato, fase_macro, fase, atividade, responsavel,
                    inicio_previsto, termino_previsto, conclusao_pct, critico,
                    dependencia, observacoes, created_at, updated_at, created_by
  - hub_auditoria_imgs: id, contrato, categoria, url, legenda, data_captura, autor, created_at
  - hub_timeline: id, contrato, tipo, titulo, descricao, autor, created_at, mencoes (jsonb)
"""
import logging
from typing import Any, Dict, List
import reflex as rx

from bomtempo.core.supabase_client import sb_select, sb_insert, sb_update, sb_delete
from bomtempo.core.audit_logger import audit_log, AuditCategory

logger = logging.getLogger(__name__)

from datetime import timezone, timedelta
_BRT = timezone(timedelta(hours=-3))


def _utc_to_brt(ts: str) -> str:
    """Convert ISO UTC timestamp → BRT (UTC-3), formatted as DD/MM/YYYY HH:MM."""
    if not ts:
        return ""
    try:
        from datetime import datetime
        dt = datetime.fromisoformat(ts.replace("Z", "+00:00")[:32])
        brt = dt.astimezone(_BRT)
        return brt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return ts[:16].replace("T", " ")


def _utc_date_to_br(ts: str) -> str:
    """Convert ISO date string YYYY-MM-DD → DD/MM/YYYY."""
    if not ts:
        return ""
    try:
        parts = ts[:10].split("-")
        if len(parts) == 3:
            return f"{parts[2]}/{parts[1]}/{parts[0]}"
    except Exception:
        pass
    return ts[:10]

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

ENTRY_TYPES = ["Atualização", "Marco", "Reunião", "Decisão", "Alerta", "Falha", "Documento", "Custo"]

AUDIT_CATEGORIES = [
    {"slug": "equipe",       "label": "Equipe com EPI",    "icon": "hard-hat",     "color": "#22c55e"},
    {"slug": "falhas",       "label": "Falhas & Logs",     "icon": "alert-triangle","color": "#EF4444"},
    {"slug": "ferramentas",  "label": "Ferramentas",       "icon": "wrench",       "color": "#2A9D8F"},
    {"slug": "gerais",       "label": "Imagens Gerais",    "icon": "image",        "color": "#C98B2A"},
]

FASE_COLORS: Dict[str, str] = {
    "civil":       "#C98B2A",
    "elétrica":    "#3B82F6",
    "eletrica":    "#3B82F6",
    "hidráulica":  "#2A9D8F",
    "hidraulica":  "#2A9D8F",
    "estrutural":  "#E89845",
    "mecânica":    "#A855F7",
    "mecanica":    "#A855F7",
    "licenciamento": "#64748B",
    "aprovações":  "#64748B",
    "aprovacoes":  "#64748B",
}


def _fase_color(fase: str) -> str:
    return FASE_COLORS.get(fase.lower().strip(), "#889999")


_DIAS_MAP = {"seg": 0, "ter": 1, "qua": 2, "qui": 3, "sex": 4, "sab": 5, "sáb": 5, "dom": 6}


def _parse_dias_uteis(dias_str: str) -> set:
    """Convert 'seg,ter,qua,qui,sex' to a set of weekday ints (0=Mon…6=Sun).
    Falls back to Mon-Fri if the string is empty or unrecognized."""
    if not dias_str:
        return {0, 1, 2, 3, 4}
    result = set()
    for d in dias_str.split(","):
        key = d.strip().lower()
        if key in _DIAS_MAP:
            result.add(_DIAS_MAP[key])
    return result if result else {0, 1, 2, 3, 4}


def _add_working_days(start_iso: str, days: int, working_days: set = None) -> str:
    """Return ISO date string for an activity that starts on start_iso and lasts `days` working days.

    The start day itself counts as day 1. So dias=1 → termino=inicio; dias=2 → termino=next working day.
    working_days: set of weekday ints (0=Mon…6=Sun). Defaults to Mon-Fri.
    """
    from datetime import date, timedelta
    if working_days is None:
        working_days = {0, 1, 2, 3, 4}
    try:
        current = date.fromisoformat(start_iso[:10])
        if not working_days or days <= 0:
            return start_iso
        # Day 1 is the start date itself; advance (days-1) additional working days
        added = 1
        while added < days:
            current += timedelta(days=1)
            if current.weekday() in working_days:
                added += 1
        return current.isoformat()
    except Exception:
        return start_iso


def _norm_str(v: object, fallback: str = "") -> str:
    if v is None or str(v) in ("None", "NaT", "nan", ""):
        return fallback
    return str(v)


def _recalc_macro_dates(macro_id: str, contrato: str, client_id: str) -> None:
    """After saving a micro, update its parent macro's inicio/termino from children dates."""
    try:
        children = sb_select("hub_atividades", filters={"parent_id": macro_id, "contrato": contrato}, limit=200)
        starts = [r["inicio_previsto"] for r in children if r.get("inicio_previsto")]
        ends = [r["termino_previsto"] for r in children if r.get("termino_previsto")]
        if starts and ends:
            sb_update("hub_atividades", filters={"id": macro_id}, data={
                "inicio_previsto": min(starts),
                "termino_previsto": max(ends),
            })
    except Exception as e:
        logger.warning(f"_recalc_macro_dates error: {e}")


def _log_schedule_diff_async(
    contrato: str,
    atividade_id: str,
    atividade_nome: str,
    old_row: dict,
    new_row: dict,
    autor: str,
    client_id: str,
) -> None:
    """Fire-and-forget: record a full diff of changed fields into hub_cronograma_log
    + summary entry in hub_timeline with AI impact note (best-effort).

    Tracked fields:
      inicio_previsto, termino_previsto, conclusao_pct, responsavel,
      peso_pct, critico, nivel, fase_macro, fase, observacoes,
      total_qty, unidade, dias_planejados
    """
    import threading

    FIELD_LABELS = {
        "inicio_previsto":  "Início",
        "termino_previsto": "Término",
        "conclusao_pct":    "Conclusão %",
        "responsavel":      "Responsável",
        "peso_pct":         "Peso %",
        "critico":          "Crítico",
        "nivel":            "Nível",
        "fase_macro":       "Fase Macro",
        "fase":             "Fase",
        "observacoes":      "Observações",
        "total_qty":        "Qtd Total",
        "unidade":          "Unidade",
        "dias_planejados":  "Dias Planejados",
        "status_atividade": "Status",
        "tipo_medicao":     "Tipo Medição",
    }

    def _fmt(field: str, v) -> str:
        v = str(v or "")
        if field in ("inicio_previsto", "termino_previsto", "data_inicio_real", "data_fim_real", "data_fim_prevista"):
            return _utc_date_to_br(v) or "—"
        if field == "critico":
            return "Sim" if v.upper() in ("TRUE", "1", "SIM", "YES") else "Não"
        return v or "—"

    def _run():
        try:
            diffs = []
            for field in FIELD_LABELS:
                old_v = str(old_row.get(field, "") or "")
                new_v = str(new_row.get(field, "") or "")
                if old_v != new_v:
                    diffs.append((field, old_v, new_v))

            if not diffs:
                return

            # Insert one log row per changed field
            for field, old_v, new_v in diffs:
                try:
                    sb_insert("hub_cronograma_log", {
                        "contrato":       contrato,
                        "atividade_id":   atividade_id or None,
                        "atividade_nome": atividade_nome[:120],
                        "campo":          field,
                        "valor_anterior": _fmt(field, old_v)[:500],
                        "valor_novo":     _fmt(field, new_v)[:500],
                        "autor":          autor or "sistema",
                        "client_id":      client_id or None,
                    })
                except Exception:
                    pass

            # Human-readable summary for timeline
            date_fields = {"inicio_previsto", "termino_previsto"}
            date_diffs = [(f, o, n) for f, o, n in diffs if f in date_fields]
            other_diffs = [(f, o, n) for f, o, n in diffs if f not in date_fields]

            parts = []
            for f, o, n in date_diffs:
                parts.append(f"{FIELD_LABELS[f]}: {_fmt(f, o)} → {_fmt(f, n)}")
            for f, o, n in other_diffs[:4]:  # cap at 4 non-date changes per entry
                label = FIELD_LABELS.get(f, f)
                parts.append(f"{label}: {_fmt(f, o)} → {_fmt(f, n)}")

            change_summary = " | ".join(parts)
            titulo = f"[Cronograma] {atividade_nome[:60]} — {len(diffs)} campo(s) alterado(s)"

            tl_id = None
            try:
                result = sb_insert("hub_timeline", {
                    "contrato":    contrato,
                    "tipo":        "Atualização",
                    "titulo":      titulo,
                    "descricao":   change_summary,
                    "autor":       autor or "sistema",
                    "mencoes":     [],
                    "is_document": False,
                    "is_cost":     False,
                    "client_id":   client_id or None,
                })
                # sb_insert returns a dict (single row) or None
                if result and isinstance(result, dict) and result.get("id"):
                    tl_id = result["id"]
                elif result and isinstance(result, list) and result[0].get("id"):
                    tl_id = result[0]["id"]
            except Exception:
                pass

            # AI impact analysis (best-effort, enriches the timeline entry)
            try:
                from bomtempo.core.ai_client import ai_client
                diff_text = "\n".join(
                    f"- {FIELD_LABELS.get(f, f)}: '{_fmt(f, o)}' → '{_fmt(f, n)}'"
                    for f, o, n in diffs
                )
                msg = [{"role": "user", "content": (
                    f"Atividade de obra: '{atividade_nome}'\n"
                    f"Alterações registradas:\n{diff_text}\n\n"
                    f"Em 1-2 frases objetivas: qual o impacto dessas mudanças no cronograma? "
                    f"Considere dependências, prazo contratual e caminho crítico."
                )}]
                ai_note = ai_client.query(msg)
                if ai_note:
                    # Update timeline entry with AI note
                    if tl_id:
                        sb_update("hub_timeline", filters={"id": tl_id},
                                  data={"descricao": f"{change_summary}\n\n🤖 {ai_note}"})
                    # Also store impacto in the log rows for this batch
                    try:
                        from datetime import datetime, timezone
                        # Update the most recent log row for this atividade
                        sb_update(
                            "hub_cronograma_log",
                            filters={"atividade_id": atividade_id, "autor": autor},
                            data={"ai_impacto": ai_note[:1000]},
                        )
                    except Exception:
                        pass
            except Exception:
                pass

        except Exception as ex:
            logger.warning(f"_log_schedule_diff_async error: {ex}")

    threading.Thread(target=_run, daemon=True).start()


# Keep legacy name as thin wrapper for backward compat
def _log_schedule_change_async(
    contrato: str, atividade: str,
    old_inicio: str, new_inicio: str,
    old_termino: str, new_termino: str,
    autor: str, client_id: str,
) -> None:
    """Legacy wrapper — delegates to _log_schedule_diff_async."""
    _log_schedule_diff_async(
        contrato=contrato,
        atividade_id="",
        atividade_nome=atividade,
        old_row={"inicio_previsto": old_inicio, "termino_previsto": old_termino},
        new_row={"inicio_previsto": new_inicio, "termino_previsto": new_termino},
        autor=autor,
        client_id=client_id,
    )


def _norm_pct(v: object) -> str:
    try:
        return str(int(float(v or 0)))
    except (ValueError, TypeError):
        return "0"


# ─────────────────────────────────────────────────────────────────────────────
# State
# ─────────────────────────────────────────────────────────────────────────────

class HubState(rx.State):
    """Hub de Operações — Cronograma, Auditoria, Timeline state."""

    # ── Loading flags ────────────────────────────────────────────────────────
    cron_loading: bool = False
    audit_loading: bool = False
    timeline_loading: bool = False

    # ══════════════════════════════════════════════════════════════════════════
    # CRONOGRAMA
    # ══════════════════════════════════════════════════════════════════════════

    # List of normalized activity dicts for the selected contract
    # Keys: id, contrato, fase_macro, fase, atividade, responsavel,
    #       inicio_previsto, termino_previsto, conclusao_pct, critico,
    #       dependencia, observacoes, color
    cron_rows: List[Dict[str, str]] = []

    # Working days config for the current project (loaded from contratos.dias_uteis_semana)
    cron_working_days_str: str = "seg,ter,qua,qui,sex"
    # Current contract code (set by load_cronograma so upload handlers can read it)
    cron_contrato: str = ""

    # Filter
    cron_fase_filter: str = ""
    cron_search: str = ""
    cron_search_input: str = ""  # UI-only: updated on_change, committed on_blur/Enter
    cron_show_only_critical: bool = False

    # Inline edit dialog
    cron_show_dialog: bool = False
    cron_edit_id: str = ""          # empty = new
    cron_pending_review_id: str = ""  # set when opening dialog for pending approval
    cron_edit_atividade: str = ""
    cron_edit_fase_macro: str = ""
    cron_edit_fase: str = ""
    cron_edit_responsavel: str = ""
    cron_edit_inicio: str = ""
    cron_edit_termino: str = ""
    cron_edit_pct: str = "0"
    cron_edit_critico: bool = False
    cron_edit_dependencia: str = ""   # legacy text name (kept for compat)
    cron_edit_dependencia_id: str = ""  # UUID of dependency activity
    cron_edit_observacoes: str = ""
    cron_saving: bool = False
    cron_error: str = ""
    # Quantity tracking (#17)
    cron_edit_total_qty: str = ""    # total planned quantity (e.g. "1456")
    cron_edit_unidade: str = ""      # unit (e.g. "perfurações", "m²", "un")
    cron_edit_dias_planejados: str = ""  # working days → auto-fill termino
    # Forecast / status fields
    cron_edit_status_atividade: str = "nao_iniciada"
    cron_edit_tipo_medicao: str = "quantidade"

    # Delete confirm
    cron_delete_id: str = ""
    cron_delete_name: str = ""
    cron_show_delete: bool = False

    # IA import from Excel/PDF
    cron_import_loading: bool = False
    cron_import_error: str = ""
    cron_import_preview: List[Dict[str, str]] = []  # proposed activities to review
    cron_import_show: bool = False
    cron_import_selected: List[str] = []  # ids of proposals to confirm

    # ══════════════════════════════════════════════════════════════════════════
    # AUDITORIA (photo gallery bolsões)
    # ══════════════════════════════════════════════════════════════════════════

    # List of all images for selected contract
    # Keys: id, contrato, categoria, url, legenda, data_captura, autor
    audit_images: List[Dict[str, str]] = []

    # Currently open bolsão slug (e.g. "equipe", "falhas", "ferramentas", "gerais")
    audit_open_category: str = ""

    # Lightbox
    audit_lightbox_url: str = ""
    audit_lightbox_legenda: str = ""
    audit_lightbox_data: str = ""
    audit_lightbox_autor: str = ""

    # Upload dialog
    audit_show_upload: bool = False
    audit_upload_category: str = ""
    audit_upload_url: str = ""
    audit_upload_legenda: str = ""
    audit_uploading: bool = False
    audit_upload_error: str = ""

    # ══════════════════════════════════════════════════════════════════════════
    # TIMELINE
    # ══════════════════════════════════════════════════════════════════════════

    # Feed of log entries for selected contract
    # Keys: id, contrato, tipo, titulo, descricao, autor, created_at,
    #       is_document, is_cost, custo_valor, custo_categoria,
    #       anexo_url, anexo_nome
    timeline_entries: List[Dict[str, str]] = []

    # New entry form
    tl_entry_type: str = "Atualização"
    tl_titulo: str = ""
    tl_descricao: str = ""
    tl_submitting: bool = False
    tl_error: str = ""

    # mention users
    tl_mention_users: List[str] = []   # usernames disponíveis para @mention

    # custo fields
    tl_custo_valor: str = ""       # valor do custo (string para input)
    tl_custo_categoria: str = "Operacional"  # categoria do custo

    # New: file attachment
    tl_anexo_url: str = ""         # URL do arquivo no Supabase Storage após upload
    tl_anexo_nome: str = ""        # nome original do arquivo
    tl_uploading_anexo: bool = False

    # Filter + search
    tl_filter_tipo: str = ""
    tl_search: str = ""            # busca por título/descrição
    tl_search_input: str = ""  # UI-only input buffer

    # ══════════════════════════════════════════════════════════════════════════
    # MACRO/MICRO hierarchy
    # ══════════════════════════════════════════════════════════════════════════

    # Edit fields for hierarchical properties
    cron_edit_nivel: str = "macro"       # "macro" | "micro"
    cron_edit_peso: str = "100"          # peso_pct relative to parent (macro) or project (macro)
    cron_edit_parent_id: str = ""        # parent macro id (only for micros)

    # Which macro rows are expanded (showing micros) — list of macro ids
    cron_expanded_macros: List[str] = []

    # Pending activities awaiting approval (role=gestor)
    pending_activities: List[Dict[str, str]] = []
    cron_approve_loading: bool = False

    # ══════════════════════════════════════════════════════════════════════════
    # GANTT PREMIUM — weather + IA climate
    # ══════════════════════════════════════════════════════════════════════════

    # IA climate analysis result
    cron_climate_analysis: str = ""
    cron_climate_loading: bool = False

    # ──────────────────────────────────────────────────────────────────────────
    # Internal helpers
    # ──────────────────────────────────────────────────────────────────────────

    @rx.var
    def filtered_cron_rows(self) -> List[Dict[str, str]]:
        """Apply fase filter, search, and critical-only to cron_rows."""
        rows = self.cron_rows
        if self.cron_fase_filter:
            rows = [r for r in rows if r.get("fase_macro", "") == self.cron_fase_filter]
        if self.cron_show_only_critical:
            rows = [r for r in rows if r.get("critico", "") == "1"]
        if self.cron_search:
            q = self.cron_search.lower()
            rows = [
                r for r in rows
                if q in r.get("atividade", "").lower()
                or q in r.get("responsavel", "").lower()
                or q in r.get("fase", "").lower()
            ]
        return rows

    @rx.var
    def cron_unique_fases(self) -> List[str]:
        seen = []
        for r in self.cron_rows:
            f = r.get("fase_macro", "").strip()
            if f and f not in seen:
                seen.append(f)
        return seen

    @rx.var
    def cron_stats(self) -> Dict[str, str]:
        total = len(self.cron_rows)
        done = sum(1 for r in self.cron_rows if int(r.get("conclusao_pct", "0") or "0") >= 100)
        critical = sum(1 for r in self.cron_rows if r.get("critico", "") == "1")
        pct = round(done / total * 100) if total else 0
        return {
            "total": str(total),
            "done": str(done),
            "critical": str(critical),
            "pct": str(pct),
        }

    @rx.var
    def audit_category_counts(self) -> Dict[str, str]:
        counts: Dict[str, str] = {}
        for cat in AUDIT_CATEGORIES:
            slug = cat["slug"]
            counts[slug] = str(sum(1 for img in self.audit_images if img.get("categoria") == slug))
        return counts

    @rx.var
    def audit_open_images(self) -> List[Dict[str, str]]:
        if not self.audit_open_category:
            return []
        return [img for img in self.audit_images if img.get("categoria") == self.audit_open_category]

    @rx.var
    def filtered_timeline(self) -> List[Dict[str, str]]:
        rows = self.timeline_entries
        if self.tl_filter_tipo:
            rows = [e for e in rows if e.get("tipo") == self.tl_filter_tipo]
        if self.tl_search:
            q = self.tl_search.lower()
            rows = [e for e in rows if q in e.get("titulo", "").lower() or q in e.get("descricao", "").lower()]
        return rows

    @rx.var
    def audit_lightbox_open(self) -> bool:
        return self.audit_lightbox_url != ""

    @rx.var
    def cron_activity_options(self) -> List[str]:
        """List of existing activity names for dependency dropdown (excludes current edit)."""
        return [
            r.get("atividade", "")
            for r in self.cron_rows
            if r.get("atividade") and r.get("id") != self.cron_edit_id
        ]

    @rx.var
    def gantt_rows(self) -> List[Dict[str, str]]:
        """
        Returns filtered_cron_rows enriched with Gantt positioning data:
          - gantt_left_pct: CSS left% within the Gantt timeline (string, e.g. "12.5")
          - gantt_width_pct: CSS width% within the Gantt timeline (string, e.g. "30.0")
          - gantt_overdue: "1" if termino_iso < today and conclusao_pct < 100
        All activities without valid ISO dates get left="0", width="100".
        """
        from datetime import date
        rows = self.filtered_cron_rows
        if not rows:
            return []

        # Compute global min/max from all rows (not just filtered) for consistent scale
        all_rows = self.cron_rows
        dates_start = []
        dates_end = []
        for r in all_rows:
            s = r.get("inicio_iso", "")
            e = r.get("termino_iso", "")
            if s and len(s) >= 10:
                try:
                    dates_start.append(date.fromisoformat(s[:10]))
                except Exception:
                    pass
            if e and len(e) >= 10:
                try:
                    dates_end.append(date.fromisoformat(e[:10]))
                except Exception:
                    pass

        if not dates_start or not dates_end:
            return [dict(r, gantt_left_pct="0", gantt_width_pct="100", gantt_overdue="0",
                         gantt_forecast_left="", gantt_forecast_width="") for r in rows]

        global_start = min(dates_start)
        global_end = max(dates_end)
        total_days = (global_end - global_start).days or 1
        today = date.today()

        result = []
        for r in rows:
            s_iso = r.get("inicio_iso", "")
            e_iso = r.get("termino_iso", "")
            try:
                s_date = date.fromisoformat(s_iso[:10]) if s_iso and len(s_iso) >= 10 else global_start
                e_date = date.fromisoformat(e_iso[:10]) if e_iso and len(e_iso) >= 10 else global_end
                left_days = (s_date - global_start).days
                dur_days = max((e_date - s_date).days, 1)
                left_pct = round(left_days / total_days * 100, 1)
                width_pct = round(dur_days / total_days * 100, 1)
                # Clamp
                if left_pct < 0:
                    left_pct = 0.0
                if left_pct + width_pct > 100:
                    width_pct = 100.0 - left_pct
                overdue = "1" if e_date < today and int(r.get("conclusao_pct", "0") or "0") < 100 else "0"
            except Exception:
                left_pct, width_pct, overdue = 0.0, 100.0, "0"
                s_date = global_start
                e_date = global_end

            # Forecast bar: from s_date to EAC (data_fim_prevista) — only for micros with qty
            gantt_forecast_left = ""
            gantt_forecast_width = ""
            try:
                total_qty_f = float(r.get("total_qty", "0") or "0")
                exec_qty_f  = float(r.get("exec_qty", "0") or "0")
                dias_plan_f = int(r.get("dias_planejados", "0") or "0")
                if total_qty_f > 0 and exec_qty_f > 0 and dias_plan_f > 0:
                    prod_plan_f = total_qty_f / dias_plan_f
                    d_inicio_f  = date.fromisoformat(s_iso[:10]) if s_iso and len(s_iso) >= 10 else None
                    if d_inicio_f:
                        dias_dec = max(1, int((today - d_inicio_f).days * 5 / 7))
                        prod_real_f = exec_qty_f / dias_dec
                        if prod_real_f > 0:
                            saldo_f = max(0.0, total_qty_f - exec_qty_f)
                            cal_rest = int((saldo_f / prod_real_f) * 1.4)
                            from datetime import timedelta
                            eac_date = today + timedelta(days=cal_rest)
                            eac_left = round((s_date - global_start).days / total_days * 100, 1)
                            eac_dur  = max((eac_date - s_date).days, 1)
                            eac_w    = round(eac_dur / total_days * 100, 1)
                            eac_left = max(0.0, eac_left)
                            if eac_left + eac_w > 100:
                                eac_w = 100.0 - eac_left
                            gantt_forecast_left  = str(eac_left)
                            gantt_forecast_width = str(max(eac_w, 0.5))
            except Exception:
                pass

            result.append(dict(
                r,
                gantt_left_pct=str(left_pct),
                gantt_width_pct=str(max(width_pct, 0.8)),
                gantt_overdue=overdue,
                gantt_forecast_left=gantt_forecast_left,
                gantt_forecast_width=gantt_forecast_width,
            ))
        return result

    @rx.var
    def gantt_date_range(self) -> Dict[str, str]:
        """Returns {'start': 'DD/MM/YYYY', 'end': 'DD/MM/YYYY'} for display."""
        from datetime import date
        if not self.cron_rows:
            return {"start": "—", "end": "—"}
        dates_s, dates_e = [], []
        for r in self.cron_rows:
            s = r.get("inicio_iso", "")
            e = r.get("termino_iso", "")
            if s and len(s) >= 10:
                try:
                    dates_s.append(date.fromisoformat(s[:10]))
                except Exception:
                    pass
            if e and len(e) >= 10:
                try:
                    dates_e.append(date.fromisoformat(e[:10]))
                except Exception:
                    pass
        if not dates_s or not dates_e:
            return {"start": "—", "end": "—"}
        def fmt(d: date) -> str:
            return d.strftime("%d/%m/%Y")
        return {"start": fmt(min(dates_s)), "end": fmt(max(dates_e))}

    @rx.var
    def cron_macro_rows(self) -> List[Dict[str, str]]:
        """All macro-level activities (nivel=='macro' or nivel missing/empty)."""
        return [r for r in self.filtered_cron_rows if r.get("nivel", "macro") in ("macro", "")]

    @rx.var
    def cron_display_rows(self) -> List[Dict[str, str]]:
        """
        Flat ordered list for rx.foreach — macros interleaved with their micros.
        Each row has _display_mode: 'macro' | 'micro' | 'pending_micro'.
        Micros only appear when their parent macro is in cron_expanded_macros.
        """
        result: List[Dict[str, str]] = []
        all_rows = self.filtered_cron_rows
        # Separate macros and micros
        macros = [r for r in all_rows if r.get("nivel", "macro") in ("macro", "")]
        micros = [r for r in all_rows if r.get("nivel", "") == "micro"]
        expanded = self.cron_expanded_macros

        for macro in macros:
            macro_id = macro.get("id", "")
            # Calculate macro_progress from micros if any
            children = [m for m in micros if m.get("parent_id", "") == macro_id]
            if children:
                total_peso = sum(int(c.get("peso_pct", "0") or "0") for c in children)
                if total_peso > 0:
                    weighted_pct = sum(
                        int(c.get("conclusao_pct", "0") or "0") * int(c.get("peso_pct", "0") or "0")
                        for c in children
                    ) / total_peso
                else:
                    weighted_pct = 0.0
                computed_pct = str(round(weighted_pct))
                has_micros = "1"
            else:
                computed_pct = macro.get("conclusao_pct", "0")
                has_micros = "0"

            is_expanded = "1" if macro_id in expanded else "0"
            result.append(dict(
                macro,
                _display_mode="macro",
                _has_micros=has_micros,
                _is_expanded=is_expanded,
                _computed_pct=computed_pct,
                _micro_count=str(len(children)),
            ))

            # Append micros if expanded
            if macro_id in expanded:
                for micro in children:
                    result.append(dict(
                        micro,
                        _display_mode="micro",
                        _has_micros="0",
                        _is_expanded="0",
                        _computed_pct=micro.get("conclusao_pct", "0"),
                        _micro_count="0",
                    ))

        # Standalone entries that are micros with no parent in current filter
        macro_ids = {m.get("id", "") for m in macros}
        orphan_micros = [m for m in micros if m.get("parent_id", "") not in macro_ids]
        for micro in orphan_micros:
            result.append(dict(
                micro,
                _display_mode="micro",
                _has_micros="0",
                _is_expanded="0",
                _computed_pct=micro.get("conclusao_pct", "0"),
                _micro_count="0",
            ))

        return result

    @rx.var
    def cron_pending_rows(self) -> List[Dict[str, str]]:
        """Activities flagged as pendente_aprovacao=True."""
        return [r for r in self.cron_rows if r.get("pendente_aprovacao", "0") == "1"]

    @rx.var
    def cron_parent_options(self) -> List[Dict[str, str]]:
        """List of macros for parent dropdown when creating/editing a micro."""
        return [
            {"id": r.get("id", ""), "label": r.get("atividade", "")}
            for r in self.cron_rows
            if r.get("nivel", "macro") in ("macro", "") and r.get("id") != self.cron_edit_id
        ]

    # ── Forecast / Produtividade computed vars ────────────────────────────────

    @rx.var
    def cron_forecast_rows(self) -> List[Dict[str, str]]:
        """
        Enriches micro activities with forecast fields:
          prod_planejada_dia, prod_real_media, desvio_pct, data_fim_prevista,
          desvio_dias, tendencia ('acima'|'dentro'|'abaixo'|'sem_dados')
        Only micros with total_qty > 0 AND exec_qty > 0 get a meaningful forecast.
        """
        from datetime import date, timedelta
        today = date.today()
        result = []
        for r in self.cron_rows:
            if r.get("nivel", "macro") != "micro":
                continue
            try:
                total_qty  = float(r.get("total_qty",  "0") or "0")
                exec_qty   = float(r.get("exec_qty",   "0") or "0")
                dias_plan  = int(r.get("dias_planejados", "0") or "0")
                pct        = int(r.get("conclusao_pct", "0") or "0")
                inicio_iso = r.get("inicio_iso", "")
                termino_iso = r.get("termino_iso", "")
            except (ValueError, TypeError):
                continue

            # produtividade planejada
            prod_plan = 0.0
            if total_qty > 0 and dias_plan > 0:
                prod_plan = total_qty / dias_plan

            # dias decorridos desde início (aproximado, ignorando feriados)
            dias_decorridos = 0
            if inicio_iso and len(inicio_iso) >= 10:
                try:
                    d_inicio = date.fromisoformat(inicio_iso[:10])
                    if d_inicio <= today:
                        dias_decorridos = max(0, (today - d_inicio).days)
                except Exception:
                    pass

            # produtividade real média (exec_qty / dias decorridos úteis estimados)
            dias_uteis_decorridos = max(1, int(dias_decorridos * 5 / 7))  # rough weekday estimate
            prod_real = exec_qty / dias_uteis_decorridos if dias_uteis_decorridos > 0 and exec_qty > 0 else 0.0

            # desvio de produtividade
            desvio_pct = 0.0
            if prod_plan > 0:
                desvio_pct = round((prod_real - prod_plan) / prod_plan * 100, 1)

            # tendência
            tol = 10.0  # ±10% tolerance
            if exec_qty == 0 or dias_decorridos < 3:
                tendencia = "sem_dados"
            elif desvio_pct >= tol:
                tendencia = "acima"
            elif desvio_pct <= -tol:
                tendencia = "abaixo"
            else:
                tendencia = "dentro"

            # data fim prevista (EAC)
            data_fim_prev_str = ""
            desvio_dias = 0
            if prod_real > 0 and total_qty > 0 and pct < 100:
                saldo = total_qty - exec_qty
                dias_restantes = max(0, int(saldo / prod_real))
                # Convert working days → calendar days (rough ×1.4)
                cal_restantes = int(dias_restantes * 1.4)
                fim_prev = today + timedelta(days=cal_restantes)
                data_fim_prev_str = fim_prev.isoformat()
                if termino_iso and len(termino_iso) >= 10:
                    try:
                        fim_plan = date.fromisoformat(termino_iso[:10])
                        desvio_dias = (fim_prev - fim_plan).days
                    except Exception:
                        pass
            elif pct >= 100:
                tendencia = "concluida"

            result.append(dict(
                r,
                _prod_planejada=f"{prod_plan:.1f}",
                _prod_real=f"{prod_real:.1f}",
                _desvio_pct=f"{desvio_pct:+.1f}",
                _tendencia=tendencia,
                _data_fim_prevista=_utc_date_to_br(data_fim_prev_str) if data_fim_prev_str else "—",
                _desvio_dias=str(desvio_dias),
                _saldo_qty=f"{max(0.0, float(r.get('total_qty','0') or '0') - float(r.get('exec_qty','0') or '0')):.1f}",
            ))
        return result

    @rx.var
    def cron_kpi_dashboard(self) -> Dict[str, str]:
        """
        KPIs de alto nível do cronograma para o dashboard previsto vs realizado:
          - pct_fisico_programado_hoje: % planejado até hoje (baseado em atividades vencidas)
          - pct_fisico_realizado: média ponderada real de todas as atividades
          - desvio_pp: diferença em pontos percentuais
          - atividades_em_risco: count de micros com tendência 'abaixo' e >5% de desvio
          - atividades_atrasadas: micros com termino < hoje e pct < 100
          - atividades_adiantadas: micros com tendencia 'acima'
          - producao_total_prevista: soma total_qty de todos os micros
          - producao_total_realizada: soma exec_qty de todos os micros
        """
        from datetime import date
        today = date.today()

        micros = [r for r in self.cron_rows if r.get("nivel", "") == "micro"]
        if not micros:
            return {
                "pct_fisico_programado_hoje": "0",
                "pct_fisico_realizado": "0",
                "desvio_pp": "0",
                "atividades_em_risco": "0",
                "atividades_atrasadas": "0",
                "atividades_adiantadas": "0",
                "producao_total_prevista": "0",
                "producao_total_realizada": "0",
                "total_micros": "0",
            }

        # Progresso realizado ponderado por peso
        total_peso = sum(int(r.get("peso_pct", "0") or "0") for r in micros)
        if total_peso > 0:
            pct_realizado = sum(
                int(r.get("conclusao_pct", "0") or "0") * int(r.get("peso_pct", "0") or "0")
                for r in micros
            ) / total_peso
        else:
            pct_realizado = sum(int(r.get("conclusao_pct", "0") or "0") for r in micros) / len(micros)

        # Programado até hoje: atividades que deveriam estar concluídas (termino <= hoje)
        vencidas = []
        for r in micros:
            t = r.get("termino_iso", "")
            if t and len(t) >= 10:
                try:
                    if date.fromisoformat(t[:10]) <= today:
                        vencidas.append(r)
                except Exception:
                    pass
        pct_programado = (len(vencidas) / len(micros) * 100) if micros else 0.0
        desvio_pp = round(pct_realizado - pct_programado, 1)

        # Atividades atrasadas: termino < hoje e pct < 100
        atrasadas = 0
        for r in micros:
            t = r.get("termino_iso", "")
            pct = int(r.get("conclusao_pct", "0") or "0")
            if t and len(t) >= 10 and pct < 100:
                try:
                    if date.fromisoformat(t[:10]) < today:
                        atrasadas += 1
                except Exception:
                    pass

        # Produção física total
        prod_prev = sum(float(r.get("total_qty", "0") or "0") for r in micros)
        prod_real = sum(float(r.get("exec_qty", "0") or "0") for r in micros)

        # Atividades em risco (baseado em forecast_rows)
        em_risco = sum(1 for r in self.cron_forecast_rows if r.get("_tendencia") == "abaixo")
        adiantadas = sum(1 for r in self.cron_forecast_rows if r.get("_tendencia") == "acima")

        return {
            "pct_fisico_programado_hoje": str(round(pct_programado, 1)),
            "pct_fisico_realizado": str(round(pct_realizado, 1)),
            "desvio_pp": f"{desvio_pp:+.1f}",
            "atividades_em_risco": str(em_risco),
            "atividades_atrasadas": str(atrasadas),
            "atividades_adiantadas": str(adiantadas),
            "producao_total_prevista": str(round(prod_prev, 1)),
            "producao_total_realizada": str(round(prod_real, 1)),
            "total_micros": str(len(micros)),
        }

    # ══════════════════════════════════════════════════════════════════════════
    # CRONOGRAMA — Load & CRUD
    # ══════════════════════════════════════════════════════════════════════════

    @rx.event(background=True)
    async def load_cronograma(self, contrato: str):
        """Load activities for a given contract from hub_atividades table."""
        async with self:
            self.cron_loading = True
            self.cron_rows = []
            self.cron_fase_filter = ""
            self.cron_search = ""

        try:
            # Load project working days config
            contrato_rows = sb_select("contratos", filters={"contrato": contrato}, limit=1)
            dias_uteis_str = "seg,ter,qua,qui,sex"
            if contrato_rows:
                dias_uteis_str = str(contrato_rows[0].get("dias_uteis_semana", "") or "seg,ter,qua,qui,sex")

            rows = sb_select(
                "hub_atividades",
                filters={"contrato": contrato},
                order="fase_macro.asc,inicio_previsto.asc",
                limit=500,
            )
            normalized = []
            for r in rows:
                fase = _norm_str(r.get("fase", ""))
                pendente_raw = r.get("pendente_aprovacao", False)
                pendente = "1" if str(pendente_raw or "").upper() in ("TRUE", "1", "SIM", "YES") else "0"
                normalized.append({
                    "id":              _norm_str(r.get("id")),
                    "contrato":        _norm_str(r.get("contrato")),
                    "fase_macro":      _norm_str(r.get("fase_macro")),
                    "fase":            fase,
                    "atividade":       _norm_str(r.get("atividade")),
                    "responsavel":     _norm_str(r.get("responsavel"), "—"),
                    "inicio_previsto": _utc_date_to_br(_norm_str(r.get("inicio_previsto"))),
                    "termino_previsto": _utc_date_to_br(_norm_str(r.get("termino_previsto"))),
                    # raw ISO dates kept for Gantt/weather lookups
                    "inicio_iso": _norm_str(r.get("inicio_previsto"))[:10],
                    "termino_iso": _norm_str(r.get("termino_previsto"))[:10],
                    "conclusao_pct":   _norm_pct(r.get("conclusao_pct")),
                    "critico":         "1" if str(r.get("critico", "") or "").upper() in ("TRUE", "1", "SIM") else "0",
                    "dependencia":     _norm_str(r.get("dependencia")),
                    "observacoes":     _norm_str(r.get("observacoes")),
                    "color":           _fase_color(fase),
                    # Hierarchical fields
                    "nivel":           _norm_str(r.get("nivel"), "macro"),
                    "parent_id":       _norm_str(r.get("parent_id")),
                    "peso_pct":        _norm_pct(r.get("peso_pct") if r.get("peso_pct") is not None else 100),
                    "pendente_aprovacao": pendente,
                    # Quantity tracking
                    "total_qty":       _norm_str(r.get("total_qty", "0") or "0"),
                    "exec_qty":        _norm_str(r.get("exec_qty", "0") or "0"),
                    "unidade":         _norm_str(r.get("unidade", "")),
                    "dias_planejados": _norm_str(r.get("dias_planejados", "0") or "0"),
                    "dependencia_id":  _norm_str(r.get("dependencia_id", "")),
                    # Forecast fields (new)
                    "status_atividade": _norm_str(r.get("status_atividade", "nao_iniciada") or "nao_iniciada"),
                    "tipo_medicao":     _norm_str(r.get("tipo_medicao", "quantidade") or "quantidade"),
                    "frente_servico":   _norm_str(r.get("frente_servico", "")),
                    "data_inicio_real": _utc_date_to_br(_norm_str(r.get("data_inicio_real") or "")),
                    "data_fim_real":    _utc_date_to_br(_norm_str(r.get("data_fim_real") or "")),
                    "data_fim_prevista": _utc_date_to_br(_norm_str(r.get("data_fim_prevista") or "")),
                    "data_inicio_real_iso": _norm_str(r.get("data_inicio_real") or "")[:10],
                    "data_fim_real_iso":    _norm_str(r.get("data_fim_real") or "")[:10],
                })
        except Exception as e:
            logger.error(f"load_cronograma error: {e}")
            normalized = []

        async with self:
            self.cron_rows = normalized
            self.cron_working_days_str = dias_uteis_str
            self.cron_contrato = contrato
            self.cron_loading = False

    def set_cron_fase_filter(self, value: str):
        self.cron_fase_filter = "" if self.cron_fase_filter == value else value

    def set_cron_search(self, value: str):
        self.cron_search = value

    def commit_cron_search(self, _value: str = ""):
        """Commit search from blur or Enter key — only then triggers filtered_cron_rows recalc."""
        self.cron_search = self.cron_search_input

    def set_cron_search_input(self, value: str):
        """Update local input var without triggering filtered_cron_rows recalc."""
        self.cron_search_input = value

    def handle_cron_search_key(self, key: str):
        if key == "Enter":
            self.cron_search = self.cron_search_input

    def toggle_cron_critical(self):
        self.cron_show_only_critical = not self.cron_show_only_critical

    @rx.event(background=True)
    async def recalculate_cron_dates(self):
        """Recalculate termino_previsto for all activities that have inicio_previsto + dias_planejados,
        using the current project's working days config. Useful after changing dias_uteis_semana."""
        from bomtempo.core.supabase_client import sb_update

        contrato = ""
        working_days_str = "seg,ter,qua,qui,sex"
        rows_snapshot = []
        async with self:
            contrato = self.cron_contrato
            working_days_str = self.cron_working_days_str
            rows_snapshot = list(self.cron_rows)

        if not contrato or not rows_snapshot:
            yield rx.toast.warning("Nenhuma atividade carregada.", duration=3000)
            return

        wd = _parse_dias_uteis(working_days_str)
        updated = 0
        for r in rows_snapshot:
            inicio_iso = r.get("inicio_iso", "")
            dias_raw = r.get("dias_planejados", "0")
            row_id = r.get("id", "")
            if not inicio_iso or not row_id:
                continue
            try:
                dias = int(dias_raw or 0)
            except (ValueError, TypeError):
                dias = 0
            if dias <= 0:
                continue
            new_termino = _add_working_days(inicio_iso, dias, wd)
            try:
                sb_update("hub_atividades", {"id": row_id}, {"termino_previsto": new_termino})
                updated += 1
            except Exception as ex:
                logger.warning(f"recalculate_cron_dates: erro ao atualizar {row_id}: {ex}")

        if updated:
            yield rx.toast.success(f"{updated} datas de término recalculadas.", duration=4000)
            yield HubState.load_cronograma(contrato)
        else:
            yield rx.toast.info("Nenhuma atividade com início + dias para recalcular.", duration=3000)

    # ── Dialog open/close ─────────────────────────────────────────────────────

    def open_cron_new_root(self):
        self.open_cron_new("")

    def open_cron_new(self, parent_id: str = ""):
        self.cron_edit_id = ""
        self.cron_edit_atividade = ""
        self.cron_edit_fase_macro = ""
        self.cron_edit_fase = ""
        self.cron_edit_responsavel = ""
        self.cron_edit_inicio = ""
        self.cron_edit_termino = ""
        self.cron_edit_pct = "0"
        self.cron_edit_critico = False
        self.cron_edit_dependencia = ""
        self.cron_edit_dependencia_id = ""
        self.cron_edit_observacoes = ""
        self.cron_edit_total_qty = ""
        self.cron_edit_unidade = ""
        self.cron_edit_dias_planejados = ""
        self.cron_edit_status_atividade = "nao_iniciada"
        self.cron_edit_tipo_medicao = "quantidade"
        self.cron_error = ""
        # Hierarchy defaults
        if parent_id:
            self.cron_edit_nivel = "micro"
            self.cron_edit_parent_id = parent_id
            # Inherit fase_macro from parent
            parent = next((r for r in self.cron_rows if r["id"] == parent_id), None)
            if parent:
                self.cron_edit_fase_macro = parent.get("fase_macro", "")
                self.cron_edit_fase = parent.get("fase", "")
        else:
            self.cron_edit_nivel = "macro"
            self.cron_edit_parent_id = ""
        self.cron_edit_peso = "100"
        self.cron_show_dialog = True

    def open_cron_edit(self, row_id: str):
        row = next((r for r in self.cron_rows if r["id"] == row_id), None)
        if not row:
            return
        self.cron_edit_id = row_id
        self.cron_edit_atividade = row.get("atividade", "")
        self.cron_edit_fase_macro = row.get("fase_macro", "")
        self.cron_edit_fase = row.get("fase", "")
        self.cron_edit_responsavel = row.get("responsavel", "")
        self.cron_edit_inicio = row.get("inicio_iso", row.get("inicio_previsto", ""))
        self.cron_edit_termino = row.get("termino_iso", row.get("termino_previsto", ""))
        self.cron_edit_pct = row.get("conclusao_pct", "0")
        self.cron_edit_critico = row.get("critico", "0") == "1"
        self.cron_edit_dependencia = row.get("dependencia", "")
        self.cron_edit_dependencia_id = row.get("dependencia_id", "")
        self.cron_edit_observacoes = row.get("observacoes", "")
        self.cron_edit_total_qty = row.get("total_qty", "")
        self.cron_edit_unidade = row.get("unidade", "")
        self.cron_edit_dias_planejados = row.get("dias_planejados", "")
        self.cron_edit_status_atividade = row.get("status_atividade", "nao_iniciada") or "nao_iniciada"
        self.cron_edit_tipo_medicao = row.get("tipo_medicao", "quantidade") or "quantidade"
        # Hierarchy
        self.cron_edit_nivel = row.get("nivel", "macro")
        self.cron_edit_parent_id = row.get("parent_id", "")
        self.cron_edit_peso = str(row.get("peso_pct", "100") or "100")
        self.cron_error = ""
        self.cron_show_dialog = True

    def open_pending_review(self, row_id: str):
        """Open the edit dialog pre-filled for a pending-approval activity."""
        self.open_cron_edit(row_id)
        self.cron_pending_review_id = row_id

    def close_cron_dialog(self):
        self.cron_show_dialog = False
        self.cron_pending_review_id = ""

    def set_cron_show_dialog(self, v: bool):
        self.cron_show_dialog = v
        if not v:
            self.cron_pending_review_id = ""

    def set_cron_edit_atividade(self, v: str): self.cron_edit_atividade = v
    def set_cron_edit_fase_macro(self, v: str):
        self.cron_edit_fase_macro = v
        # Para atividades macro, o nome é a própria fase macro
        if self.cron_edit_nivel == "macro":
            self.cron_edit_atividade = v
    def set_cron_edit_nivel(self, v: str):
        self.cron_edit_nivel = v
        # Ao mudar para macro, sincroniza o nome com fase_macro
        if v == "macro" and self.cron_edit_fase_macro.strip():
            self.cron_edit_atividade = self.cron_edit_fase_macro
    def set_cron_edit_fase(self, v: str): self.cron_edit_fase = v
    def set_cron_edit_responsavel(self, v: str): self.cron_edit_responsavel = v
    def set_cron_edit_inicio(self, v: str):
        self.cron_edit_inicio = v
        # Auto-recalculate termino if dias_planejados is set
        if self.cron_edit_dias_planejados.strip():
            try:
                dias = int(self.cron_edit_dias_planejados.strip())
                wd = _parse_dias_uteis(self.cron_working_days_str)
                self.cron_edit_termino = _add_working_days(v, dias, wd)
            except Exception:
                pass
    def set_cron_edit_termino(self, v: str): self.cron_edit_termino = v
    def set_cron_edit_pct(self, v): self.cron_edit_pct = str(v)
    def toggle_cron_edit_critico(self): self.cron_edit_critico = not self.cron_edit_critico
    def set_cron_edit_dependencia(self, v: str): self.cron_edit_dependencia = "" if v == "__none__" else v
    def set_cron_edit_dependencia_id(self, dep_id: str):
        """Select a dependency by activity id — auto-fill inicio from its termino."""
        if dep_id in ("", "__none__"):
            self.cron_edit_dependencia_id = ""
            self.cron_edit_dependencia = ""
            return
        self.cron_edit_dependencia_id = dep_id
        dep_row = next((r for r in self.cron_rows if r["id"] == dep_id), None)
        if dep_row:
            self.cron_edit_dependencia = dep_row.get("atividade", "")
            dep_termino = dep_row.get("termino_iso", "")
            if dep_termino:
                self.cron_edit_inicio = dep_termino
                # Recalculate termino if dias_planejados is set
                if self.cron_edit_dias_planejados.strip():
                    try:
                        dias = int(self.cron_edit_dias_planejados.strip())
                        wd = _parse_dias_uteis(self.cron_working_days_str)
                        self.cron_edit_termino = _add_working_days(dep_termino, dias, wd)
                    except Exception:
                        pass
    def set_cron_edit_dias_planejados(self, v):
        v = str(v) if v is not None else ""
        self.cron_edit_dias_planejados = v
        # Auto-calculate termino when dias changes and inicio is set
        if v.strip() and self.cron_edit_inicio:
            try:
                dias = int(v.strip())
                wd = _parse_dias_uteis(self.cron_working_days_str)
                self.cron_edit_termino = _add_working_days(self.cron_edit_inicio, dias, wd)
            except Exception:
                pass
    def set_cron_edit_total_qty(self, v): self.cron_edit_total_qty = str(v) if v is not None else ""
    def set_cron_edit_unidade(self, v: str): self.cron_edit_unidade = v
    def set_cron_edit_observacoes(self, v: str): self.cron_edit_observacoes = v
    def set_cron_edit_status_atividade(self, v: str): self.cron_edit_status_atividade = v
    def set_cron_edit_tipo_medicao(self, v: str): self.cron_edit_tipo_medicao = v
    def set_cron_edit_peso(self, v): self.cron_edit_peso = str(v) if v is not None else "100"
    def set_cron_edit_parent_id(self, v: str): self.cron_edit_parent_id = v

    def toggle_macro_expanded(self, macro_id: str):
        if macro_id in self.cron_expanded_macros:
            self.cron_expanded_macros = [x for x in self.cron_expanded_macros if x != macro_id]
        else:
            new_list = list(self.cron_expanded_macros)
            new_list.append(macro_id)
            self.cron_expanded_macros = new_list

    # ── Save ─────────────────────────────────────────────────────────────────

    @rx.event(background=True)
    async def save_cron_activity(self):
        """INSERT or UPDATE a hub_atividade row."""
        from bomtempo.state.global_state import GlobalState

        contrato = ""
        atividade_nome = ""
        edit_id = ""
        old_inicio = ""
        old_termino = ""
        edit_inicio = ""
        edit_termino = ""
        edit_pct = 0
        edit_critico = False
        edit_dependencia = ""
        edit_dependencia_id = ""
        edit_observacoes = ""
        edit_fase_macro = ""
        edit_fase = ""
        edit_responsavel = ""
        edit_nivel = "macro"
        edit_parent_id = ""
        edit_peso = 100
        edit_total_qty = 0.0
        edit_unidade = ""
        edit_dias_planejados = 0
        edit_status_atividade = "nao_iniciada"
        edit_tipo_medicao = "quantidade"
        pending_review_id = ""
        username = ""

        async with self:
            if not self.cron_edit_atividade.strip():
                self.cron_error = "Nome da atividade é obrigatório."
                return
            self.cron_saving = True
            self.cron_error = ""
            contrato = self.cron_rows[0].get("contrato", "") if self.cron_rows else ""
            atividade_nome = self.cron_edit_atividade.strip()
            edit_id = self.cron_edit_id
            pending_review_id = self.cron_pending_review_id
            edit_fase_macro = self.cron_edit_fase_macro.strip()
            edit_fase = self.cron_edit_fase.strip()
            edit_responsavel = self.cron_edit_responsavel.strip()
            edit_inicio = self.cron_edit_inicio or None
            edit_termino = self.cron_edit_termino or None
            edit_pct = int(self.cron_edit_pct or 0)
            edit_critico = self.cron_edit_critico
            edit_dependencia = self.cron_edit_dependencia.strip()
            edit_dependencia_id = self.cron_edit_dependencia_id.strip() or None
            edit_observacoes = self.cron_edit_observacoes.strip()
            edit_nivel = self.cron_edit_nivel or "macro"
            edit_parent_id = self.cron_edit_parent_id or None
            edit_peso = int(self.cron_edit_peso or 100)
            # #12 — Warn if sibling micros sum > 100%
            if edit_nivel == "micro" and edit_parent_id:
                siblings = [
                    r for r in self.cron_rows
                    if r.get("parent_id") == edit_parent_id
                    and r.get("nivel") == "micro"
                    and r.get("id") != (self.cron_edit_id or "")
                ]
                sibling_sum = sum(int(r.get("peso_pct", "0") or "0") for r in siblings)
                if sibling_sum + edit_peso > 100:
                    self.cron_error = (
                        f"⚠ Pesos das sub-atividades somam {sibling_sum + edit_peso}% "
                        f"(máximo 100%). Ajuste o peso para ≤ {100 - sibling_sum}%."
                    )
                    self.cron_saving = False
                    return
            try:
                edit_total_qty = float(self.cron_edit_total_qty.replace(",", ".")) if self.cron_edit_total_qty.strip() else 0.0
            except Exception:
                edit_total_qty = 0.0
            edit_unidade = self.cron_edit_unidade.strip()
            try:
                edit_dias_planejados = int(self.cron_edit_dias_planejados.strip()) if self.cron_edit_dias_planejados.strip() else 0
            except Exception:
                edit_dias_planejados = 0
            edit_status_atividade = self.cron_edit_status_atividade or "nao_iniciada"
            edit_tipo_medicao = self.cron_edit_tipo_medicao or "quantidade"
            # Capture old values for full diff log
            old_snapshot: Dict[str, str] = {}
            if edit_id:
                old_row = next((r for r in self.cron_rows if r["id"] == edit_id), {})
                old_inicio = old_row.get("inicio_iso", "")
                old_termino = old_row.get("termino_iso", "")
                old_snapshot = {
                    "inicio_previsto":  old_inicio,
                    "termino_previsto": old_termino,
                    "conclusao_pct":    old_row.get("conclusao_pct", ""),
                    "responsavel":      old_row.get("responsavel", ""),
                    "peso_pct":         old_row.get("peso_pct", ""),
                    "critico":          old_row.get("critico", ""),
                    "nivel":            old_row.get("nivel", ""),
                    "fase_macro":       old_row.get("fase_macro", ""),
                    "fase":             old_row.get("fase", ""),
                    "observacoes":      old_row.get("observacoes", ""),
                    "total_qty":        old_row.get("total_qty", ""),
                    "unidade":          old_row.get("unidade", ""),
                    "dias_planejados":  old_row.get("dias_planejados", ""),
                    "status_atividade": old_row.get("status_atividade", ""),
                    "tipo_medicao":     old_row.get("tipo_medicao", ""),
                }
            # Sempre busca GlobalState para client_id + fallback de contrato
            gs = await self.get_state(GlobalState)
            if not contrato:
                contrato = str(gs.selected_contrato or gs.selected_project or "")
            client_id = str(gs.current_client_id or "")
            username = str(gs.current_user_name or "")

        try:
            data: Dict[str, Any] = {
                "contrato":          contrato,
                "fase_macro":        edit_fase_macro,
                "fase":              edit_fase,
                "atividade":         atividade_nome,
                "responsavel":       edit_responsavel,
                "inicio_previsto":   edit_inicio,
                "termino_previsto":  edit_termino,
                "conclusao_pct":     edit_pct,
                "critico":           edit_critico,
                "dependencia":       edit_dependencia,
                "dependencia_id":    edit_dependencia_id,
                "observacoes":       edit_observacoes,
                "nivel":             edit_nivel,
                "parent_id":         edit_parent_id,
                "peso_pct":          edit_peso,
                "total_qty":         edit_total_qty,
                "unidade":           edit_unidade,
                "dias_planejados":   edit_dias_planejados,
                "status_atividade":  edit_status_atividade,
                "tipo_medicao":      edit_tipo_medicao,
                "client_id":         client_id,
            }

            if edit_id:
                if pending_review_id:
                    data["pendente_aprovacao"] = False
                sb_update("hub_atividades", filters={"id": edit_id}, data=data)
                action = f"Atividade '{atividade_nome}' {'aprovada' if pending_review_id else 'atualizada'}"

                # ── Full diff log → hub_cronograma_log + hub_timeline ─────────
                if not pending_review_id:
                    new_snapshot = {
                        "inicio_previsto":  edit_inicio or "",
                        "termino_previsto": edit_termino or "",
                        "conclusao_pct":    str(edit_pct),
                        "responsavel":      edit_responsavel,
                        "peso_pct":         str(edit_peso),
                        "critico":          "1" if edit_critico else "0",
                        "nivel":            edit_nivel,
                        "fase_macro":       edit_fase_macro,
                        "fase":             edit_fase,
                        "observacoes":      edit_observacoes,
                        "total_qty":        str(edit_total_qty),
                        "unidade":          edit_unidade,
                        "dias_planejados":  str(edit_dias_planejados),
                        "status_atividade": edit_status_atividade,
                        "tipo_medicao":     edit_tipo_medicao,
                    }
                    _log_schedule_diff_async(
                        contrato=contrato,
                        atividade_id=edit_id,
                        atividade_nome=atividade_nome,
                        old_row=old_snapshot,
                        new_row=new_snapshot,
                        autor=username,
                        client_id=client_id,
                    )
            else:
                sb_insert("hub_atividades", data)
                action = f"Atividade '{atividade_nome}' criada"

            # ── Auto-derive macro dates from micros ───────────────────────────
            if edit_nivel == "micro" and edit_parent_id:
                _recalc_macro_dates(edit_parent_id, contrato, client_id)

            audit_log(
                category=AuditCategory.DATA_EDIT,
                action=action,
                username=username,
                entity_type="hub_atividades",
                entity_id=edit_id or "new",
                metadata={"contrato": contrato, "atividade": atividade_nome},
            )

        except Exception as e:
            logger.error(f"save_cron_activity error: {e}")
            async with self:
                self.cron_error = f"Erro: {str(e)[:120]}"
                self.cron_saving = False
            return

        # Reload
        async with self:
            self.cron_show_dialog = False
            self.cron_saving = False
            self.cron_pending_review_id = ""

        yield HubState.load_cronograma(contrato)

    # ── Delete ────────────────────────────────────────────────────────────────

    def request_cron_delete(self, row_id: str):
        row = next((r for r in self.cron_rows if r["id"] == row_id), None)
        self.cron_delete_id = row_id
        self.cron_delete_name = row.get("atividade", row_id) if row else row_id
        self.cron_show_delete = True

    def cancel_cron_delete(self):
        self.cron_delete_id = ""
        self.cron_show_delete = False

    @rx.event(background=True)
    async def confirm_cron_delete(self):
        row_id = ""
        name = ""
        contrato = ""
        async with self:
            row_id = str(self.cron_delete_id)
            name = str(self.cron_delete_name)
            contrato = self.cron_rows[0].get("contrato", "") if self.cron_rows else ""
            self.cron_show_delete = False
            self.cron_delete_id = ""
            # Optimistic UI: remove imediatamente da lista local antes do DB call
            self.cron_rows = [r for r in self.cron_rows if r.get("id") != row_id]

        if not row_id:
            logger.error("confirm_cron_delete: row_id vazio, abortando.")
            yield rx.toast.error("ID da atividade não encontrado.", duration=4000)
            return

        delete_ok = False
        try:
            logger.info(f"confirm_cron_delete: deleting id={row_id!r} name={name!r}")
            # Retry once on transient failure
            delete_ok = sb_delete("hub_atividades", filters={"id": row_id})
            if not delete_ok:
                import time as _time
                _time.sleep(0.5)
                delete_ok = sb_delete("hub_atividades", filters={"id": row_id})
            if delete_ok:
                audit_log(
                    category=AuditCategory.DATA_DELETE,
                    action=f"Atividade '{name}' excluída",
                    username="",
                    entity_type="hub_atividades",
                    entity_id=row_id,
                )
            else:
                logger.error(f"confirm_cron_delete: sb_delete retornou False para id={row_id}")
                yield rx.toast.error("Erro ao excluir atividade no banco de dados.", duration=4000)
        except Exception as e:
            logger.error(f"confirm_cron_delete error: {e}")
            yield rx.toast.error(f"Erro ao excluir: {str(e)[:100]}", duration=4000)

        if contrato:
            yield HubState.load_cronograma(contrato)

    # ── Approve / Reject pending activities ───────────────────────────────────

    @rx.event(background=True)
    async def approve_pending_activity(self, activity_id: str):
        """Approve a pending activity (set pendente_aprovacao=False)."""
        contrato = ""
        async with self:
            self.cron_approve_loading = True
            contrato = self.cron_rows[0].get("contrato", "") if self.cron_rows else ""
        try:
            sb_update("hub_atividades", filters={"id": activity_id}, data={"pendente_aprovacao": False})
        except Exception as e:
            logger.error(f"approve_pending_activity error: {e}")
        async with self:
            self.cron_approve_loading = False
        if contrato:
            yield HubState.load_cronograma(contrato)

    @rx.event(background=True)
    async def reject_pending_activity(self, activity_id: str):
        """Reject (delete) a pending activity."""
        contrato = ""
        async with self:
            self.cron_approve_loading = True
            contrato = self.cron_rows[0].get("contrato", "") if self.cron_rows else ""
        try:
            sb_delete("hub_atividades", filters={"id": activity_id})
        except Exception as e:
            logger.error(f"reject_pending_activity error: {e}")
        async with self:
            self.cron_approve_loading = False
        if contrato:
            yield HubState.load_cronograma(contrato)

    # ── IA Climate Analysis ────────────────────────────────────────────────────

    @rx.event(background=True)
    async def analyze_climate_impact(self):
        """Cross-reference scheduled activities with weather forecast via IA."""
        from bomtempo.state.global_state import GlobalState

        weather = {}
        rows = []
        async with self:
            self.cron_climate_loading = True
            self.cron_climate_analysis = ""
            gs = await self.get_state(GlobalState)
            weather = dict(gs.weather_data) if gs.weather_data else {}
            rows = list(self.cron_rows)

        # Build context
        from datetime import date
        today_iso = date.today().isoformat()
        weather_summary = ""
        if weather and weather.get("daily_time"):
            lines = []
            for i, dt in enumerate(weather["daily_time"][:7]):
                prob = weather["daily_rain_prob"][i] if i < len(weather.get("daily_rain_prob", [])) else 0
                rain = weather["daily_rain_sum"][i] if i < len(weather.get("daily_rain_sum", [])) else 0
                lines.append(f"  {dt}: chuva {rain:.1f}mm, prob. {prob}%")
            weather_summary = "\n".join(lines)
        else:
            weather_summary = "  Dados climáticos não disponíveis"

        activities_summary = ""
        if rows:
            act_lines = []
            for r in rows[:20]:
                act_lines.append(
                    f"  • {r['atividade']} ({r['fase_macro']}) | {r['inicio_previsto']} → {r['termino_previsto']} | {r['conclusao_pct']}% | crítico: {'sim' if r['critico']=='1' else 'não'}"
                )
            activities_summary = "\n".join(act_lines)
        else:
            activities_summary = "  Nenhuma atividade cadastrada"

        messages = [
            {
                "role": "user",
                "content": (
                    f"Você é um engenheiro de obras sênior analisando impacto climático em um cronograma de construção.\n\n"
                    f"DATA ATUAL: {today_iso}\n\n"
                    f"PREVISÃO CLIMÁTICA (próximos 7 dias):\n{weather_summary}\n\n"
                    f"ATIVIDADES DO CRONOGRAMA:\n{activities_summary}\n\n"
                    f"Analise quais atividades previstas para os próximos 7 dias serão impactadas pela chuva, "
                    f"especialmente as críticas. Dê recomendações práticas: o que antecipar, o que proteger, "
                    f"o que reagendar. Seja direto e objetivo, máximo 200 palavras."
                ),
            }
        ]

        try:
            import queue as _queue
            import threading
            result_queue: _queue.Queue = _queue.Queue()

            def _run_analysis():
                try:
                    # Use direct AI call
                    from bomtempo.core.ai_client import ai_client
                    full_text = ""
                    for chunk in ai_client.query_stream(messages):
                        full_text += chunk
                    result_queue.put(("ok", full_text))
                except Exception as ex:
                    result_queue.put(("err", str(ex)))

            t = threading.Thread(target=_run_analysis, daemon=True)
            t.start()
            t.join(timeout=60)

            if not result_queue.empty():
                status, text = result_queue.get()
                async with self:
                    self.cron_climate_analysis = text if status == "ok" else f"Erro na análise: {text[:200]}"
                if status == "ok":
                    yield rx.toast.success("Análise climática concluída — veja o resultado abaixo.", duration=5000)
                    yield rx.call_script("document.getElementById('climate-analysis-panel')?.scrollIntoView({behavior:'smooth',block:'start'})")
            else:
                async with self:
                    self.cron_climate_analysis = "Tempo esgotado ao consultar IA. Tente novamente."
                yield rx.toast.warning("Análise climática: tempo esgotado.", duration=4000)

        except Exception as e:
            logger.error(f"analyze_climate_impact error: {e}")
            async with self:
                self.cron_climate_analysis = f"Erro: {str(e)[:200]}"
        finally:
            async with self:
                self.cron_climate_loading = False

    def clear_climate_analysis(self):
        self.cron_climate_analysis = ""

    # ── IA Import from Excel / PDF ─────────────────────────────────────────────

    async def import_cronograma_ia(self, files: list[rx.UploadFile]):
        """Receive uploaded file via rx.upload, parse content, send to IA, propose activities."""
        from bomtempo.state.global_state import GlobalState
        import json

        # Regular async upload handler — state writes are direct (no async with self needed)
        self.cron_import_loading = True
        self.cron_import_error = ""
        self.cron_import_preview = []
        # Use cron_contrato (set by load_cronograma) as the authoritative source
        contrato = self.cron_contrato or (self.cron_rows[0].get("contrato", "") if self.cron_rows else "")
        yield  # flush loading=True to UI

        file_text = ""
        if files:
            upload_file = files[0]
            name = upload_file.filename
            raw_bytes = await upload_file.read()
            try:
                ext = name.rsplit(".", 1)[-1].lower() if "." in name else ""
                if ext in ("xlsx", "xls"):
                    import io
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
                    all_lines = []
                    # Itera TODAS as abas do Excel
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        all_lines.append(f"\n=== ABA: {sheet_name} ===")
                        for row in ws.iter_rows(values_only=True):
                            cells = [str(c) if c is not None else "" for c in row]
                            row_text = " | ".join(cells).strip(" |")
                            if row_text:  # ignora linhas completamente vazias
                                all_lines.append(row_text)
                    file_text = "\n".join(all_lines[:400])
                elif ext == "csv":
                    file_text = raw_bytes.decode("utf-8", errors="replace")[:8000]
                else:
                    # PDF or other: decode as text best-effort
                    file_text = raw_bytes.decode("utf-8", errors="replace")[:8000]
            except Exception as ex:
                self.cron_import_error = f"Erro ao ler arquivo: {ex}"
                self.cron_import_loading = False
                return

        if not file_text.strip():
            self.cron_import_error = "Arquivo vazio ou não legível."
            self.cron_import_loading = False
            return

        prompt = f"""Você é um assistente especialista em gestão de obras e cronogramas. Analise o arquivo abaixo e extraia EXATAMENTE as atividades que estão explicitamente listadas. NÃO invente, NÃO duplique, NÃO crie atividades que não estejam no documento.

REGRAS ESTRITAS:
1. Extraia APENAS atividades que estão EXPLICITAMENTE no arquivo. Se não encontrar, retorne lista vazia.
2. NÃO crie atividades derivadas ou inferidas. Se uma linha descreve quantidades de uma atividade já listada, use esses dados para preencher total_qty/unidade da atividade existente — NÃO crie outra atividade.
3. NÃO duplique atividades: se a mesma atividade aparece em abas diferentes (ex: aba "Cronograma" e aba "Atividades"), use a ocorrência que contiver mais informações (datas, quantidades), ignorando as demais.
4. Para hierarquia: linhas de fase/disciplina principal → nivel="macro"; sub-atividades → nivel="micro".
5. Datas: se o arquivo tiver colunas de datas, extraia-as. Se for um Gantt (barras em colunas de datas), leia o cabeçalho para determinar inicio_previsto e termino_previsto. Datas em DD/MM/AAAA, DD/MM/AA → converta para YYYY-MM-DD. Se não houver data, deixe vazio — NÃO invente datas.
6. dias_planejados: calcule como dias úteis entre inicio e termino SE ambas as datas existirem. Se houver coluna de duração explícita, use-a. Se não houver nenhuma informação, deixe 0.
7. total_qty e unidade: extraia apenas se houver coluna explícita com quantidade (ex: "500 m²", "120 kg"). Se não houver, deixe 0 e vazio.
8. Atividades marcadas como críticas, em vermelho ou com asterisco → critico=true.
9. Se houver coluna de responsável/equipe, extraia em "responsavel".
10. Ao final, adicione um campo "_confidence" (0.0 a 1.0) indicando sua confiança na extração:
    - 1.0: arquivo estruturado com datas e atividades claras
    - 0.7-0.9: bom estrutura mas algumas ambiguidades
    - 0.4-0.6: pouca estrutura, muita inferência necessária
    - <0.4: arquivo difícil de interpretar, resultados incertos

ARQUIVO:
{file_text[:7000]}

Retorne SOMENTE um JSON válido com dois campos: "activities" (array) e "confidence" (número 0.0-1.0), sem texto antes ou depois, sem markdown:
{{
  "confidence": 0.9,
  "activities": [
    {{
      "atividade": "Nome exato da atividade conforme o arquivo",
      "fase_macro": "Fase/disciplina principal",
      "fase": "Sub-fase se houver, senão vazio",
      "responsavel": "Responsável/equipe se disponível, senão vazio",
      "inicio_previsto": "YYYY-MM-DD ou vazio",
      "termino_previsto": "YYYY-MM-DD ou vazio",
      "dias_planejados": número inteiro de dias úteis ou 0,
      "total_qty": número ou 0,
      "unidade": "m, m², m³, kg, und, kWh, kW, kWp ou vazio",
      "critico": true ou false,
      "nivel": "macro" ou "micro",
      "observacoes": "notas relevantes ou vazio"
    }}
  ]
}}"""

        import asyncio, queue as _queue, threading
        result_q: _queue.Queue = _queue.Queue()

        def _call_ai():
            try:
                from bomtempo.core.ai_client import ai_client
                resp = ai_client.query([{"role": "user", "content": prompt}])
                result_q.put(("ok", resp))
            except Exception as ex:
                result_q.put(("err", str(ex)))

        loop = asyncio.get_event_loop()
        await loop.run_in_executor(None, _call_ai)

        if result_q.empty():
            self.cron_import_error = "IA não respondeu a tempo. Tente novamente."
            self.cron_import_loading = False
            return

        status, raw = result_q.get()
        if status == "err":
            self.cron_import_error = f"Erro IA: {raw[:200]}"
            self.cron_import_loading = False
            return

        # Parse JSON from response
        try:
            # Strip markdown code fences if present
            cleaned = raw.strip()
            if cleaned.startswith("```"):
                cleaned = cleaned.split("\n", 1)[-1].rsplit("```", 1)[0]
            parsed = json.loads(cleaned)
            # Support both new format {"confidence": X, "activities": [...]} and legacy plain list
            if isinstance(parsed, dict):
                confidence = float(parsed.get("confidence", 0.5))
                proposals = parsed.get("activities", [])
            elif isinstance(parsed, list):
                confidence = 0.5  # legacy: assume medium confidence
                proposals = parsed
            else:
                raise ValueError("Resposta não é um objeto nem lista")
            if not isinstance(proposals, list):
                raise ValueError("Campo 'activities' não é uma lista")
        except Exception as ex:
            self.cron_import_error = f"IA retornou formato inválido: {ex}"
            self.cron_import_loading = False
            return

        # Dedup by normalized activity name to prevent duplicates from multi-sheet parsing
        seen_names: dict = {}
        deduped = []
        for p in proposals:
            name_key = str(p.get("atividade", "")).strip().lower()
            if not name_key:
                continue
            if name_key not in seen_names:
                seen_names[name_key] = p
                deduped.append(p)
            else:
                # Keep whichever has more data (prefer the one with dates)
                existing = seen_names[name_key]
                existing_has_dates = bool(existing.get("inicio_previsto") or existing.get("termino_previsto"))
                new_has_dates = bool(p.get("inicio_previsto") or p.get("termino_previsto"))
                if new_has_dates and not existing_has_dates:
                    # Replace with the version that has dates
                    seen_names[name_key] = p
                    deduped = [p if x is existing else x for x in deduped]

        # Confidence label for display
        if confidence >= 0.8:
            conf_label = f"Alta ({int(confidence * 100)}%)"
        elif confidence >= 0.5:
            conf_label = f"Média ({int(confidence * 100)}%)"
        else:
            conf_label = f"Baixa ({int(confidence * 100)}%) — revise com atenção"

        # Normalize proposals into display rows with temp IDs
        import uuid
        preview = []
        for i, p in enumerate(deduped[:50]):
            preview.append({
                "_tmp_id":          str(uuid.uuid4()),
                "atividade":        str(p.get("atividade", f"Atividade {i+1}")),
                "fase_macro":       str(p.get("fase_macro", "")),
                "fase":             str(p.get("fase", "")),
                "responsavel":      str(p.get("responsavel", "")),
                "inicio_previsto":  str(p.get("inicio_previsto", "")),
                "termino_previsto": str(p.get("termino_previsto", "")),
                "dias_planejados":  str(p.get("dias_planejados", 0) or 0),
                "total_qty":        str(p.get("total_qty", 0) or 0),
                "unidade":          str(p.get("unidade", "") or ""),
                "critico":          "1" if p.get("critico") else "0",
                "nivel":            str(p.get("nivel", "macro")),
                "observacoes":      str(p.get("observacoes", "")),
                "contrato":         contrato,
            })

        # Count items with 0 dias_planejados for consistency warning
        zero_days = sum(1 for r in preview if r["dias_planejados"] == "0")
        has_dates = sum(1 for r in preview if r["inicio_previsto"] or r["termino_previsto"])

        msg = f"{len(preview)} atividades identificadas. Confiança: {conf_label}."
        if zero_days > 0 and has_dates == 0:
            msg += f" Atenção: {zero_days} atividade(s) sem datas — preencha manualmente após importar."
        elif zero_days > 0:
            msg += f" {zero_days} atividade(s) sem duração calculada."

        self.cron_import_preview = preview
        self.cron_import_selected = [r["_tmp_id"] for r in preview]
        self.cron_import_show = True
        self.cron_import_loading = False
        yield rx.toast.success(msg, duration=8000)

    def toggle_import_activity(self, tmp_id: str):
        if tmp_id in self.cron_import_selected:
            self.cron_import_selected = [x for x in self.cron_import_selected if x != tmp_id]
        else:
            new = list(self.cron_import_selected)
            new.append(tmp_id)
            self.cron_import_selected = new

    def select_all_import(self):
        self.cron_import_selected = [r["_tmp_id"] for r in self.cron_import_preview]

    def deselect_all_import(self):
        self.cron_import_selected = []

    def close_import_preview(self):
        self.cron_import_show = False
        self.cron_import_preview = []
        self.cron_import_selected = []

    @rx.event(background=True)
    async def confirm_import_cronograma(self):
        """Bulk-insert selected proposals into hub_atividades."""
        from bomtempo.state.global_state import GlobalState

        client_id = ""
        selected_ids = set()
        to_insert = []
        contrato = ""

        working_days_str = "seg,ter,qua,qui,sex"
        async with self:
            gs = await self.get_state(GlobalState)
            client_id = str(gs.current_client_id or "")
            selected_ids = set(self.cron_import_selected)
            to_insert = [r for r in self.cron_import_preview if r["_tmp_id"] in selected_ids]
            contrato = to_insert[0]["contrato"] if to_insert else ""
            working_days_str = self.cron_working_days_str
            self.cron_import_loading = True

        if not to_insert:
            async with self:
                self.cron_import_loading = False
            yield rx.toast.warning("Nenhuma atividade selecionada.", duration=3000)
            return

        logger.info(f"confirm_import_cronograma: inserindo {len(to_insert)} atividades para contrato={contrato!r}")

        errors = 0
        inserted = 0
        for p in to_insert:
            try:
                dias = int(p.get("dias_planejados", 0) or 0)
                inicio = p.get("inicio_previsto", "") or None
                termino = p.get("termino_previsto", "") or None
                # Auto-calc termino from dias if termino missing but inicio+dias set
                if inicio and dias and not termino:
                    wd = _parse_dias_uteis(working_days_str)
                    termino = _add_working_days(inicio, dias, wd)
                total_qty = p.get("total_qty", 0) or 0
                try:
                    total_qty = float(total_qty) if total_qty else 0
                except Exception:
                    total_qty = 0
                unidade = p.get("unidade", "") or ""
                # Determina tipo_medicao com base na unidade/qty
                tipo_medicao = "quantidade" if total_qty > 0 else "percentual"
                sb_insert("hub_atividades", {
                    "contrato":          contrato,
                    "fase_macro":        p.get("fase_macro", ""),
                    "fase":              p.get("fase", ""),
                    "atividade":         p.get("atividade", ""),
                    "responsavel":       p.get("responsavel", "") or None,
                    "inicio_previsto":   inicio,
                    "termino_previsto":  termino,
                    "conclusao_pct":     0,
                    "critico":           p.get("critico", "0") == "1",
                    "nivel":             p.get("nivel", "macro"),
                    "peso_pct":          100,
                    "dias_planejados":   dias,
                    "total_qty":         total_qty,
                    "unidade":           unidade,
                    "observacoes":       p.get("observacoes", ""),
                    "status_atividade":  "nao_iniciada",
                    "tipo_medicao":      tipo_medicao,
                    "client_id":         client_id or None,
                })
                inserted += 1
            except Exception as ex:
                logger.warning(f"confirm_import error on row '{p.get('atividade','?')}': {ex}")
                errors += 1

        async with self:
            self.cron_import_show = False
            self.cron_import_preview = []
            self.cron_import_selected = []
            self.cron_import_loading = False

        if inserted:
            yield rx.toast.success(f"{inserted} atividade(s) importadas com sucesso!", duration=5000)
        if errors:
            yield rx.toast.warning(f"{errors} atividade(s) falharam na importação.", duration=5000)
        if contrato:
            yield HubState.load_cronograma(contrato)

    # ══════════════════════════════════════════════════════════════════════════
    # AUDITORIA — Load, gallery, lightbox
    # ══════════════════════════════════════════════════════════════════════════

    @rx.event(background=True)
    async def load_auditoria(self, contrato: str):
        async with self:
            self.audit_loading = True
            self.audit_images = []
            self.audit_open_category = ""

        # Captura client_id para isolamento de tenant
        client_id = ""
        try:
            from bomtempo.state.global_state import GlobalState
            _gs = await self.get_state(GlobalState)
            client_id = str(_gs.current_client_id or "")
        except Exception:
            pass

        try:
            # 1. Manual uploads in hub_auditoria_imgs
            _audit_filters: dict = {"contrato": contrato}
            if client_id:
                _audit_filters["client_id"] = client_id
            rows = sb_select(
                "hub_auditoria_imgs",
                filters=_audit_filters,
                order="created_at.desc",
                limit=500,
            )
            imgs = [
                {
                    "id":           _norm_str(r.get("id")),
                    "contrato":     _norm_str(r.get("contrato")),
                    "categoria":    _norm_str(r.get("categoria")),
                    "url":          _norm_str(r.get("url")),
                    "legenda":      _norm_str(r.get("legenda")),
                    "data_captura": _utc_date_to_br(_norm_str(r.get("data_captura"))),
                    "autor":        _norm_str(r.get("autor"), "—"),
                }
                for r in rows
            ]

            # 2. Integrate RDO photos from rdo_master (epi, ferramentas, evidencias)
            try:
                _rdo_filters: dict = {"contrato": contrato}
                if client_id:
                    _rdo_filters["client_id"] = client_id
                rdos = sb_select(
                    "rdo_master",
                    filters=_rdo_filters,
                    order="created_at.desc",
                    limit=200,
                )
                import json as _json
                for rdo in (rdos or []):
                    rdo_date = _utc_date_to_br(_norm_str(rdo.get("created_at", "") or rdo.get("data_rdo", "")))
                    rdo_autor = _norm_str(rdo.get("mestre_id", rdo.get("responsavel_tecnico", "RDO")))
                    rdo_id = _norm_str(rdo.get("id_rdo", rdo.get("id", "")))

                    # EPI photo
                    epi_url = _norm_str(rdo.get("epi_foto_url", ""))
                    if epi_url:
                        imgs.append({
                            "id":           f"rdo_epi_{rdo_id}",
                            "contrato":     contrato,
                            "categoria":    "equipe",
                            "url":          epi_url,
                            "legenda":      f"EPI — RDO {rdo_id[:8]}",
                            "data_captura": rdo_date,
                            "autor":        rdo_autor,
                        })

                    # Ferramentas photo
                    ferr_url = _norm_str(rdo.get("ferramentas_foto_url", ""))
                    if ferr_url:
                        imgs.append({
                            "id":           f"rdo_ferr_{rdo_id}",
                            "contrato":     contrato,
                            "categoria":    "ferramentas",
                            "url":          ferr_url,
                            "legenda":      f"Ferramentas — RDO {rdo_id[:8]}",
                            "data_captura": rdo_date,
                            "autor":        rdo_autor,
                        })

                    # Evidence photos (jsonb array)
                    evidencias_raw = rdo.get("evidencias") or []
                    if isinstance(evidencias_raw, str):
                        try:
                            evidencias_raw = _json.loads(evidencias_raw)
                        except Exception:
                            evidencias_raw = []
                    for idx, ev in enumerate(evidencias_raw or []):
                        ev_url = _norm_str(ev.get("foto_url", "") if isinstance(ev, dict) else "")
                        if ev_url:
                            legenda = _norm_str(ev.get("legenda", "") if isinstance(ev, dict) else "")
                            imgs.append({
                                "id":           f"rdo_ev_{rdo_id}_{idx}",
                                "contrato":     contrato,
                                "categoria":    "gerais",
                                "url":          ev_url,
                                "legenda":      legenda or f"Foto {idx+1} — RDO {rdo_id[:8]}",
                                "data_captura": rdo_date,
                                "autor":        rdo_autor,
                            })
            except Exception as e2:
                logger.warning(f"load_auditoria RDO integration error (non-fatal): {e2}")

        except Exception as e:
            logger.error(f"load_auditoria error: {e}")
            imgs = []

        async with self:
            self.audit_images = imgs
            self.audit_loading = False

    def open_audit_category(self, slug: str):
        # Toggle: clicking same category closes it
        self.audit_open_category = "" if self.audit_open_category == slug else slug

    def close_audit_category(self):
        self.audit_open_category = ""
        self.audit_lightbox_url = ""

    def open_lightbox(self, img_id: str):
        img = next((i for i in self.audit_images if i["id"] == img_id), None)
        if img:
            self.audit_lightbox_url = img["url"]
            self.audit_lightbox_legenda = img["legenda"]
            self.audit_lightbox_data = img["data_captura"]
            self.audit_lightbox_autor = img["autor"]

    def close_lightbox(self):
        self.audit_lightbox_url = ""

    def open_audit_upload(self, slug: str):
        self.audit_upload_category = slug
        self.audit_upload_url = ""
        self.audit_upload_legenda = ""
        self.audit_upload_error = ""
        self.audit_show_upload = True

    def close_audit_upload(self):
        self.audit_show_upload = False

    def set_audit_upload_url(self, v: str): self.audit_upload_url = v
    def set_audit_upload_legenda(self, v: str): self.audit_upload_legenda = v

    @rx.event(background=True)
    async def save_audit_image(self):
        from bomtempo.state.global_state import GlobalState

        contrato = ""
        autor = ""
        upload_category = ""
        upload_url = ""
        upload_legenda = ""

        async with self:
            if not self.audit_upload_url.strip():
                self.audit_upload_error = "URL da imagem é obrigatória."
                return
            self.audit_uploading = True
            self.audit_upload_error = ""
            contrato = self.audit_images[0]["contrato"] if self.audit_images else ""
            upload_category = str(self.audit_upload_category)
            upload_url = str(self.audit_upload_url).strip()
            upload_legenda = str(self.audit_upload_legenda).strip()
            gs = await self.get_state(GlobalState)
            autor = str(gs.current_user_name or "")
            if not contrato:
                contrato = str(gs.selected_contrato or gs.selected_project or "")

        from datetime import date as _date
        try:
            sb_insert("hub_auditoria_imgs", {
                "contrato":     contrato,
                "categoria":    upload_category,
                "url":          upload_url,
                "legenda":      upload_legenda,
                "autor":        autor,
                "data_captura": _date.today().isoformat(),
                "client_id":    str(gs.current_client_id or ""),
            })
        except Exception as e:
            logger.error(f"save_audit_image error: {e}")
            async with self:
                self.audit_upload_error = f"Erro: {str(e)[:100]}"
                self.audit_uploading = False
            return

        async with self:
            self.audit_show_upload = False
            self.audit_uploading = False

        yield HubState.load_auditoria(contrato)

    @rx.event(background=True)
    async def delete_audit_image(self, img_id: str):
        contrato = ""
        async with self:
            contrato = self.audit_images[0]["contrato"] if self.audit_images else ""
        try:
            sb_delete("hub_auditoria_imgs", filters={"id": img_id})
        except Exception as e:
            logger.error(f"delete_audit_image error: {e}")
        if contrato:
            yield HubState.load_auditoria(contrato)

    # ══════════════════════════════════════════════════════════════════════════
    # TIMELINE — Load & post
    # ══════════════════════════════════════════════════════════════════════════

    @rx.event(background=True)
    async def load_timeline(self, contrato: str):
        async with self:
            self.timeline_loading = True
            self.timeline_entries = []

        # Captura client_id para isolamento de tenant
        _tl_client_id = ""
        try:
            from bomtempo.state.global_state import GlobalState
            _tl_gs = await self.get_state(GlobalState)
            _tl_client_id = str(_tl_gs.current_client_id or "")
        except Exception:
            pass

        try:
            _tl_filters: dict = {"contrato": contrato}
            if _tl_client_id:
                _tl_filters["client_id"] = _tl_client_id
            rows = sb_select(
                "hub_timeline",
                filters=_tl_filters,
                order="created_at.desc",
                limit=200,
            )
            entries = [
                {
                    "id":              _norm_str(r.get("id")),
                    "contrato":        _norm_str(r.get("contrato")),
                    "tipo":            _norm_str(r.get("tipo"), "Atualização"),
                    "titulo":          _norm_str(r.get("titulo")),
                    "descricao":       _norm_str(r.get("descricao")),
                    "autor":           _norm_str(r.get("autor"), "—"),
                    "created_at":      _utc_to_brt(_norm_str(r.get("created_at"))),
                    "is_document":     "1" if r.get("is_document") else "0",
                    "is_cost":         "1" if r.get("is_cost") else "0",
                    "custo_valor":     _norm_str(r.get("custo_valor")),
                    "custo_categoria": _norm_str(r.get("custo_categoria")),
                    "anexo_url":       _norm_str(r.get("anexo_url")),
                    "anexo_nome":      _norm_str(r.get("anexo_nome")),
                }
                for r in rows
            ]
        except Exception as e:
            logger.error(f"load_timeline error: {e}")
            entries = []

        # Load mention users list — filtered by tenant
        try:
            _login_filters = {}
            if _tl_client_id:
                _login_filters["client_id"] = _tl_client_id
            login_rows = sb_select("login", filters=_login_filters if _login_filters else None, limit=100) or []
            mention_users = [str(r.get("user", "")).strip() for r in login_rows if r.get("user")]
        except Exception:
            mention_users = []

        async with self:
            self.timeline_entries = entries
            self.timeline_loading = False
            if mention_users:
                self.tl_mention_users = mention_users

    def set_tl_entry_type(self, v: str): self.tl_entry_type = v
    def set_tl_titulo(self, v: str): self.tl_titulo = v
    def set_tl_descricao(self, v: str): self.tl_descricao = v
    def set_tl_filter_tipo(self, v: str):
        self.tl_filter_tipo = "" if self.tl_filter_tipo == v else v
    def set_tl_search(self, v: str): self.tl_search = v
    def set_tl_search_input(self, v: str): self.tl_search_input = v
    def commit_tl_search(self, _v: str = ""):
        self.tl_search = self.tl_search_input
    def handle_tl_search_key(self, key: str):
        if key == "Enter":
            self.tl_search = self.tl_search_input
    def set_tl_custo_valor(self, v: str): self.tl_custo_valor = v
    def set_tl_custo_categoria(self, v: str): self.tl_custo_categoria = v

    async def upload_tl_anexo(self, files: list[rx.UploadFile]):
        """Upload file attachment to Supabase Storage bucket 'timeline-anexos'."""
        if not files:
            return
        file = files[0]
        self.tl_uploading_anexo = True
        yield

        try:
            import asyncio as _asyncio
            import re as _re
            from datetime import datetime as _dt
            from bomtempo.core.supabase_client import sb_storage_ensure_bucket, sb_storage_upload

            data = await file.read()
            nome = getattr(file, "filename", getattr(file, "name", "arquivo"))
            safe_name = _re.sub(r"[^\w\.\-]", "_", nome)
            path = f"{_dt.now().strftime('%Y%m%d_%H%M%S')}_{safe_name}"

            loop = _asyncio.get_event_loop()
            await loop.run_in_executor(None, lambda: sb_storage_ensure_bucket("timeline-anexos", public=True))
            url = await loop.run_in_executor(None, lambda: sb_storage_upload("timeline-anexos", path, data, "application/octet-stream"))

            self.tl_anexo_url = url or ""
            self.tl_anexo_nome = nome
            self.tl_uploading_anexo = False
        except Exception as e:
            logger.error(f"upload_tl_anexo error: {e}")
            self.tl_uploading_anexo = False
            self.tl_error = f"Erro no upload: {str(e)[:80]}"

    @rx.event(background=True)
    async def submit_timeline_entry(self):
        from bomtempo.state.global_state import GlobalState

        contrato = ""
        autor = ""
        entry_tipo = ""
        entry_titulo = ""
        entry_descricao = ""
        entry_is_document = False
        entry_is_cost = False
        entry_custo_valor = ""
        entry_custo_categoria = ""
        entry_anexo_url = ""
        entry_anexo_nome = ""
        entry_mencoes: list = []

        async with self:
            if not self.tl_titulo.strip():
                self.tl_error = "Título é obrigatório."
                return
            self.tl_submitting = True
            self.tl_error = ""
            contrato = self.timeline_entries[0]["contrato"] if self.timeline_entries else ""
            entry_tipo = str(self.tl_entry_type)
            entry_titulo = str(self.tl_titulo).strip()
            entry_descricao = str(self.tl_descricao).strip()
            entry_is_document = entry_tipo == "Documento"
            entry_is_cost = entry_tipo == "Custo"
            entry_custo_valor = str(self.tl_custo_valor)
            entry_custo_categoria = str(self.tl_custo_categoria)
            entry_anexo_url = str(self.tl_anexo_url)
            entry_anexo_nome = str(self.tl_anexo_nome)
            # Extract @mentions from title + description
            import re as _re
            raw_text = f"{entry_titulo} {entry_descricao}"
            entry_mencoes = list(set(_re.findall(r"@(\w+)", raw_text)))
            gs = await self.get_state(GlobalState)
            autor = str(gs.current_user_name or "")
            if not contrato:
                contrato = str(gs.selected_contrato or gs.selected_project or "")

        try:
            sb_insert("hub_timeline", {
                "contrato":        contrato,
                "tipo":            entry_tipo,
                "titulo":          entry_titulo,
                "descricao":       entry_descricao,
                "autor":           autor,
                "mencoes":         entry_mencoes,
                "is_document":     entry_is_document,
                "is_cost":         entry_is_cost,
                "custo_valor":     float(entry_custo_valor.replace(",", ".")) if entry_is_cost and entry_custo_valor else None,
                "custo_categoria": entry_custo_categoria if entry_is_cost else None,
                "anexo_url":       entry_anexo_url or None,
                "anexo_nome":      entry_anexo_nome or None,
                "client_id":       str(gs.current_client_id or ""),
            })
            audit_log(
                category=AuditCategory.DATA_EDIT,
                action=f"Timeline entry: [{entry_tipo}] {entry_titulo[:60]}",
                username=autor,
                entity_type="hub_timeline",
                entity_id=contrato,
            )
            # Create notifications for @mentioned users
            if entry_mencoes:
                _notif_msg = f"@{autor} mencionou você: [{entry_tipo}] {entry_titulo[:80]}"
                for mentioned_user in entry_mencoes:
                    if mentioned_user:
                        try:
                            sb_insert("user_notifications", {
                                "recipient": mentioned_user,
                                "sender": autor,
                                "message": _notif_msg,
                                "source_type": "mention",
                                "source_id": contrato,
                                "contrato": contrato,
                                "read": False,
                                "client_id": str(gs.current_client_id or ""),
                            })
                        except Exception as ne:
                            logger.warning(f"Failed to create notification for @{mentioned_user}: {ne}")
        except Exception as e:
            logger.error(f"submit_timeline_entry error: {e}")
            async with self:
                self.tl_error = f"Erro: {str(e)[:100]}"
                self.tl_submitting = False
            return

        async with self:
            self.tl_titulo = ""
            self.tl_descricao = ""
            self.tl_entry_type = "Atualização"
            self.tl_custo_valor = ""
            self.tl_custo_categoria = "Operacional"
            self.tl_anexo_url = ""
            self.tl_anexo_nome = ""
            self.tl_submitting = False

        yield HubState.load_timeline(contrato)

    @rx.event(background=True)
    async def delete_timeline_entry(self, entry_id: str):
        contrato = ""
        async with self:
            contrato = self.timeline_entries[0]["contrato"] if self.timeline_entries else ""
        try:
            sb_delete("hub_timeline", filters={"id": entry_id})
        except Exception as e:
            logger.error(f"delete_timeline_entry error: {e}")
        if contrato:
            yield HubState.load_timeline(contrato)
